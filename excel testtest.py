import openpyxl 
from openpyxl.styles import PatternFill
import os

g = os.walk(r"D:\1\試作無期")  
fill = PatternFill(patternType=None)
for path,dir_list,file_list in g:  
    for file_name in file_list:  
        link = os.path.join(path, file_name)
        print(link)
        wb = openpyxl.load_workbook(r"{0}".format(link))
        sheet = wb.active
        for row in sheet:
            for col in row:
                sheet[col.coordinate].fill = fill
        wb.save(r"{0}".format(link))
