#这是       已经完成 千叶11人

import openpyxl
from dateutil.relativedelta import relativedelta
mainList = openpyxl.load_workbook("D:\\1\\千葉3課.xlsx")
mainListsheet = mainList.active
listName = []
listPay = []
startDate = []

for row in range(13,86):
    for col in range(1,2):
        listName.append(mainListsheet.cell(column = col,row = row).value)

for row in range(13,86):
    for col in range(2,3):
        listPay.append(mainListsheet.cell(column = col,row = row).value)

for row in range(13,86):
    for col in range(4,5):
        value = mainListsheet.cell(column = col,row = row).value
        value = value + relativedelta(years = 3)
        startDate.append(value)
        
newCreat = openpyxl.load_workbook("D:\\1\\【2020新様式】千葉3課　盛り付け 有期.xlsx")
newCreatsheet = newCreat.active
n=12
for i in range(len(listName)):
    newCreatsheet["A2"] = listName[i]
    filename = "【2020新様式】千葉3課　有期 "+str(n)+listName[i]+".xlsx"
    newCreatsheet["V51"] = listPay[i]
    newCreatsheet["X25"] = startDate[i]
    newCreat.save("D:\\1\\試作無期\\"+filename)
    n+=1


