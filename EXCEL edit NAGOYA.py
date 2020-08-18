#这是       已经完成 千叶11人

import openpyxl
from dateutil.relativedelta import relativedelta
mainList = openpyxl.load_workbook("D:\\1\\NGY\\在籍者時給リスト　名古屋地区(上河内）00.xlsx")
mainListsheet = mainList.active
listName = []
listPay = []
startDate = []
yukiORmuki = []
department = []
leader = []

for row in range(24,77):
    for col in range(1,2):
        listName.append(mainListsheet.cell(column = col,row = row).value)

for row in range(24,77):
    for col in range(2,3):
        listPay.append(mainListsheet.cell(column = col,row = row).value)

for row in range(24,77):
    for col in range(3,4):
        yukiORmuki.append(mainListsheet.cell(column = col,row = row).value)

for row in range(24,77):
    for col in range(4,5):
        value = mainListsheet.cell(column = col,row = row).value
        value = value + relativedelta(years = 3)
        startDate.append(value)

for row in range(24,77):
    for col in range(5,6):
        department.append(mainListsheet.cell(column = col,row = row).value)

for row in range(24,77):
    for col in range(6,7):
        cell = mainListsheet.cell(column = col,row = row).value
        if cell == None:
            cell = ""
        leader.append(cell)

print(len(listPay))
print(len(listName))



fileY = openpyxl.load_workbook("D:\\1\\NGY\\【2020新様式】名古屋3課　ライン.xlsx")
Ysheet = fileY.active

fileM = openpyxl.load_workbook("D:\\1\\NGY\\【2020新様式】【無期】名古屋3課　ライン.xlsx")
Msheet = fileM.active
#有期：個人抵触日　X25　　PAY　V51　　READER　O64
n = 0
for i in (listName):
    filename = "【2020新様式】名古屋　"+str(n+24).zfill(2)+" "+yukiORmuki[n]+" "+department[n]+" "+str(listPay[n])+" "+listName[n]+" "+leader[n]+".xlsx"
    if yukiORmuki[n] == "有期":
        
        Ysheet["A2"] = listName[n]
        Ysheet["X25"] = startDate[n]
        Ysheet["V51"] = listPay[n]
        if leader[n] == "リーダー":
            Ysheet["O64"] = "リーダー手当　80円/時（基本時給に含む）"
        else:
            Ysheet["O64"] = "無"


        fileY.save("D:\\1\\NGY\\発行済\\"+filename)
    elif yukiORmuki[n] == "無期":
        Msheet["A2"] = listName[n]
        Msheet["V48"] = listPay[n]
        if leader[n] == "リーダー":
            Msheet["O61"] = "リーダー手当　80円/時（基本時給に含む）"
        else:
            Msheet["O61"] = "無"
        fileM.save("D:\\1\\NGY\\発行済\\"+filename)

    print(filename)
    n+=1
    