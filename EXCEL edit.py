#这是无期专用 已经完成 千叶11人

import openpyxl
mainList = openpyxl.load_workbook("D:\\1\\千葉3課.xlsx")
mainListsheet = mainList.active
listName = []
listPay = []
yuukiORmuki = []

for row in range(2,86):
    for col in range(1,2):
        listName.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,86):
    for col in range(2,3):
        listPay.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,86):
    for col in range(3,4):
        yuukiORmuki.append(mainListsheet.cell(column = col,row = row).value)

newCreat = openpyxl.load_workbook("D:\\1\\【2020新様式】千葉3課　盛り付け 無期.xlsx")
newCreatsheet = newCreat.active
n=1
for i in range(11):
    newCreatsheet["A2"] = listName[i]
    filename = "【2020新様式】千葉3課　無期 "+str(n)+listName[i]+".xlsx"
    newCreatsheet["V48"] = listPay[i]
    newCreat.save("D:\\1\\試作無期\\"+filename)
    n+=1