#这是神奈川3课

import openpyxl
from dateutil.relativedelta import relativedelta
mainList = openpyxl.load_workbook("D:\\1\\KA\\book1.xlsx")
mainListsheet = mainList.active
listName = []
listPay = []
yuukiORmuki = []
startDate = []

for row in range(2,67):
    for col in range(1,2):
        listName.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,67):
    for col in range(2,3):
        listPay.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,67):
    for col in range(3,4):
        yuukiORmuki.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,67):
    for col in range(4,5):
        value = mainListsheet.cell(column = col,row = row).value
        value = value + relativedelta(years = 3)
        startDate.append(value)
print(yuukiORmuki[1:10])
newCreatM = openpyxl.load_workbook("D:\\1\\KA\\【2020新様式】神奈川3課　軽食　無期.xlsx")
newCreatsheetM = newCreatM.active



newCreatY = openpyxl.load_workbook("D:\\1\\KA\\【2020新様式】神奈川3課　軽食.xlsx")
newCreatsheetY = newCreatY.active

for i in range(len(listName)): #总遍历
    if yuukiORmuki[i] == "無期":  #判断是否有期无期
        newCreatsheetM["A2"] = listName[i]
        filename = "【2020新様式】神奈川3課　無期 "+str(i).zfill(2)+listName[i]+".xlsx"
        newCreatsheetM["V48"] = listPay[i]
        newCreatM.save("D:\\1\\KA\\2020.04スタッフ分\\"+filename)
#如果是有期
    else:
        newCreatsheetY["A2"] = listName[i]
        filename = "【2020新様式】神奈川3課　有期　"+str(i).zfill(2)+listName[i]+".xlsx"
        newCreatsheetY["V51"] = listPay[i]
        newCreatsheetY["X25"] = startDate[i]
        newCreatY.save("D:\\1\\KA\\2020.04スタッフ分\\"+filename)
    
