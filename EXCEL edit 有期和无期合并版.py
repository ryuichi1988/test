#这是       已经完成 千叶11人

import openpyxl
from dateutil.relativedelta import relativedelta
mainList = openpyxl.load_workbook("D:\\1\\SN\\book.xlsx")
mainListsheet = mainList.active
listName = []
listPay = []
startDate = []
yukiORmuki = []
department = []
leader = []

for row in range(2,69):
    for col in range(1,2):
        listName.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,69):
    for col in range(2,3):
        listPay.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,69):
    for col in range(3,4):
        yukiORmuki.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,69):
    for col in range(4,5):
        value = mainListsheet.cell(column = col,row = row).value
        value = value + relativedelta(years = 3)
        startDate.append(value)

for row in range(2,69):
    for col in range(5,6):
        department.append(mainListsheet.cell(column = col,row = row).value)

for row in range(2,69):
    for col in range(6,7):
        leader.append(mainListsheet.cell(column = col,row = row).value)

#yikaOPEN
yikaFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南1課　茹で業務.xlsx")
yikasheet = yikaFileOpen.active #yikaSHIEET有期

yikaFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南1課無期　茹で業務.xlsx")
yikamuki = yikaFileOpenM.active #yikaMUKI無期
yikamuki["G20"] = "無期雇用"
yikamuki["G21"] = "　　　　　　　　　期間の定め　　　　　　無"


#生産管理OPEN
seikanFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南生産管理課　資材業務.xlsx")
seikansheet = seikanFileOpen.active #SEIKANSHIEET有期

seikanFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南生産管理課　資材業務.xlsx")
seikanmuki = seikanFileOpenM.active #SEIKANMUKI無期


#加熱OPEN
kanetsuFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南2課　加熱揚げ物業務.xlsx")
kanetsusheet = kanetsuFileOpen.active

kanetsuFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南2課無期　加熱揚げ物業務.xlsx")
kanetsumuki = kanetsuFileOpenM.active
kanetsumuki["G20"] = "無期雇用"
kanetsumuki["G21"] = "　　　　　　　　　期間の定め　　　　　　無"

#具材OPEN
guzaiFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南2課　具材業務.xlsx")
guzaisheet = guzaiFileOpen.active

guzaiFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南2課無期　具材業務.xlsx")
guzaimuki = guzaiFileOpenM.active
guzaimuki["G20"] = "無期雇用"
guzaimuki["G21"] = "　　　　　　　　　期間の定め　　　　　　無"

#3課OPEN
toppiFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南3課　うどん業務.xlsx")
toppisheet = toppiFileOpen.active

toppiFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南3課無期　うどん業務.xlsx")
toppimuki = toppiFileOpenM.active
toppimuki["G20"] = "無期雇用"
toppimuki["G21"] = "　　　　　　　　　期間の定め　　　　　　無"

#品管OPEN
hinkanFileOpen = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南品質管理課　非重量.xlsx")
hinkansheet = hinkanFileOpen.active

hinkanFileOpenM = openpyxl.load_workbook("D:\\1\\SN\\【2020新様式】湘南品質管理課　非重量.xlsx")
hinkanmuki = hinkanFileOpenM.active
hinkanmuki["G20"] = "無期雇用"
hinkanmuki["G21"] = "　　　　　　　　　期間の定め　　　　　　無"


for i in range(len(listName)):
    if department[i] == "生産管理":#生産
        if yukiORmuki[i] == "有期":    #生産有期

            seikansheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if seikansheet["AA51"].value == "円":
                seikansheet["V51"] = listPay[i]
            elif seikansheet["AA55"].value == "円":
                seikansheet["V51"] = listPay[i]
            elif seikansheet["AA59"].value == "円":
                seikansheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            seikansheet["X25"] = startDate[i]
            seikanFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#生産無期
            print("生産管理になります。このメッセージは表示されないはず")

            #生産完了

    elif department[i] == "製麺茹で":#1課
        if yukiORmuki[i] == "有期":    #1課有期

            yikasheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  yikasheet["AA56"].value == "円":
                 yikasheet["V56"] = listPay[i]
            elif  yikasheet["AA55"].value == "円":
                 yikasheet["V55"] = listPay[i]
            elif  yikasheet["AA59"].value == "円":
                 yikasheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            yikasheet["X25"] = startDate[i]
            yikaFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#1課無期
            yikamuki["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  yikamuki["AA56"].value == "円":
                 yikamuki["V56"] = listPay[i]
            elif  yikamuki["AA55"].value == "円":
                 yikamuki["V55"] = listPay[i]
            elif  yikamuki["AA59"].value == "円":
                 yikamuki["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            yikaFileOpenM.save("D:\\1\\SN\\creat\\"+filename)
            #1課完了

    elif department[i] == "品質管理":#品質管理開始
        if yukiORmuki[i] == "有期":    #品管有期

            hinkansheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  hinkansheet["AA51"].value == "円":
                 hinkansheet["V51"] = listPay[i]
            elif  hinkansheet["AA55"].value == "円":
                 hinkansheet["V55"] = listPay[i]
            elif  hinkansheet["AA59"].value == "円":
                 hinkansheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            hinkansheet["X25"] = startDate[i]
            hinkanFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#品管無期
            hinkanmuki["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  hinkanmuki["AA51"].value == "円":
                 hinkanmuki["V51"] = listPay[i]
            elif  hinkanmuki["AA55"].value == "円":
                 hinkanmuki["V55"] = listPay[i]
            elif  hinkanmuki["AA59"].value == "円":
                 hinkanmuki["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            hinkanFileOpenM.save("D:\\1\\SN\\creat\\"+filename)
            #品管完了


    elif department[i] == "具材":#具材開始
        if yukiORmuki[i] == "有期":    #具材有期

            guzaisheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  guzaisheet["AA51"].value == "円":
                 guzaisheet["V51"] = listPay[i]
            elif  guzaisheet["AA55"].value == "円":
                 guzaisheet["V55"] = listPay[i]
            elif  guzaisheet["AA59"].value == "円":
                 guzaisheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            guzaisheet["X25"] = startDate[i]
            guzaiFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#具材無期
            guzaimuki["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  guzaimuki["AA51"].value == "円":
                 guzaimuki["V51"] = listPay[i]
            elif  guzaimuki["AA55"].value == "円":
                 guzaimuki["V55"] = listPay[i]
            elif  guzaimuki["AA48"].value == "円":
                 guzaimuki["V48"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            guzaiFileOpenM.save("D:\\1\\SN\\creat\\"+filename)
            #具材完了


    elif department[i] == "加熱":#加熱開始
        if yukiORmuki[i] == "有期":    #加熱有期

            kanetsusheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  kanetsusheet["AA51"].value == "円":
                 kanetsusheet["V51"] = listPay[i]
            elif  kanetsusheet["AA55"].value == "円":
                 kanetsusheet["V55"] = listPay[i]
            elif  kanetsusheet["AA59"].value == "円":
                 kanetsusheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            kanetsusheet["X25"] = startDate[i]
            kanetsuFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#加熱無期
            kanetsumuki["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  kanetsumuki["AA51"].value == "円":
                 kanetsumuki["V51"] = listPay[i]
            elif  kanetsumuki["AA55"].value == "円":
                 kanetsumuki["V51"] = listPay[i]
            elif  kanetsumuki["AA52"].value == "円":
                 kanetsumuki["V52"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            kanetsuFileOpenM.save("D:\\1\\SN\\creat\\"+filename)
            #加熱完了


    elif department[i] == "トッピング":#トッピング開始
        if yukiORmuki[i] == "有期":    #トッピング有期

            toppisheet["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  toppisheet["AA51"].value == "円":
                 toppisheet["V51"] = listPay[i]
            elif  toppisheet["AA55"].value == "円":
                 toppisheet["V51"] = listPay[i]
            elif  toppisheet["AA59"].value == "円":
                 toppisheet["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            toppisheet["X25"] = startDate[i]
            toppiFileOpen.save("D:\\1\\SN\\creat\\"+filename)
        else:#トッピング無期
            toppimuki["A2"] = listName[i]
            filename = "【2020新様式】湘南"+str(i).zfill(2)+department[i]+yukiORmuki[i]+listName[i]+".xlsx"
            if  toppimuki["AA48"].value == "円":
                 toppimuki["V48"] = listPay[i]
            elif  toppimuki["AA55"].value == "円":
                 toppimuki["V51"] = listPay[i]
            elif  toppimuki["AA59"].value == "円":
                 toppimuki["V59"] = listPay[i]
            else:
                print("{}{}{}出现错误".format(i,department[i],listName[i]))
                break
            toppiFileOpenM.save("D:\\1\\SN\\creat\\"+filename)
            #トッピング完了



